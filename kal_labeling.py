#!/usr/bin/env python3
"""Create the Kal Labeling Master-Target and optional XLSX exports.

The script reads displayed Google Sheets values (not formulas), keeps only
valid AA groups, writes three label rows per source row, creates a
00_Datenvalidierung tab, and asks before uploading XLSX files to Drive.
"""

from __future__ import annotations

import argparse
import re
import shutil
from collections import OrderedDict
from copy import copy
from pathlib import Path
from typing import Any

SOURCE_SPREADSHEET_ID = "1c5ra-S3bzWLQ9JSoUuxpiIexW57YqhtVeNd_RWHfZa4"
SOURCE_SHEET_NAME = "KAU_PE MUs Summary"
SOURCE_SHEET_GID = 1564108399
MASTER_SPREADSHEET_ID = "1mxgUZApuu-KAwGl0HiQagQWpyAvF3vcpX9Zx1-04MbY"
TARGET_DRIVE_FOLDER_ID = "1RlXfLuGxKVBWwjx833UFqlfnEUNaCron"

COL_ASSET_ID = 4       # E
COL_DESCRIPTION = 6    # G
COL_NEW_FLO = 23       # X
COL_EQM_NUMBER = 25    # Z
COL_GROUP = 26         # AA

GROUP_PATTERN = re.compile(r"^\s*\d+\s*-\s*\S.*$")
PLACEHOLDER_PATTERN = re.compile(
    r"^(?:#(?:N/A|REF!|VALUE!|NAME\?|DIV/0!)|keine gefunden|nichts gefunden)$",
    re.IGNORECASE,
)
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def clean_text_value(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return text.lstrip("'’\ufeff")


def normalize_row(row: list[Any], width: int = 27) -> list[str]:
    values = ["" if value is None else str(value) for value in row]
    return (values + [""] * width)[:width]


def group_sort_key(name: str) -> tuple[int, str]:
    match = re.match(r"^\s*(\d+)", name)
    return (int(match.group(1)) if match else 10**9, name.casefold())


def safe_sheet_title(raw_title: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "-", raw_title).strip().strip("'")
    title = title[:100] or "Ohne Name"
    candidate = title
    counter = 2
    while candidate.casefold() in used:
        suffix = f" ~{counter}"
        candidate = f"{title[:100 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate.casefold())
    return candidate


def safe_filename(raw_title: str) -> str:
    title = re.sub(r"[\\/:*?\"<>|\x00-\x1f]", "-", raw_title).strip().rstrip(".")
    return (title[:180] or "Ohne Name") + ".xlsx"


def get_credentials(credentials_file: Path | None):
    import google.auth
    from google.oauth2 import service_account

    if credentials_file:
        return service_account.Credentials.from_service_account_file(
            credentials_file,
            scopes=SCOPES,
        )
    credentials, _ = google.auth.default(scopes=SCOPES)
    return credentials


def get_sheet_properties(sheets_api, spreadsheet_id: str) -> list[dict[str, Any]]:
    response = (
        sheets_api.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, fields="sheets.properties")
        .execute()
    )
    return [sheet["properties"] for sheet in response.get("sheets", [])]


def rgb(hex_color: str) -> dict[str, float]:
    value = hex_color.lstrip("#")
    return {
        "red": int(value[0:2], 16) / 255,
        "green": int(value[2:4], 16) / 255,
        "blue": int(value[4:6], 16) / 255,
    }


def build_label_rows(rows: list[list[str]]) -> list[list[str]]:
    result = []
    for row in rows:
        description = clean_text_value(row[COL_DESCRIPTION])
        eqm_number = clean_text_value(row[COL_EQM_NUMBER])
        new_flo = clean_text_value(row[COL_NEW_FLO])
        asset_id = clean_text_value(row[COL_ASSET_ID])
        result.extend([
            [eqm_number, "", description, "", "", ""],
            [new_flo, "", description, "", "", ""],
            [asset_id, "", "", "", "", ""],
        ])
    return result


def build_target_rows(group_name: str, rows: list[list[str]]) -> list[list[str]]:
    """Ergänzt jede Liste um den Reiternamen am Anfang und am Ende."""
    marker = [clean_text_value(group_name), "", "", "", "", ""]
    return [marker] + build_label_rows(rows) + [marker.copy()]


def load_template(template_path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    value_wb = load_workbook(template_path, data_only=True, read_only=True)
    value_ws = value_wb[value_wb.sheetnames[0]]
    headers = [
        [
            "" if value_ws.cell(row=r, column=c).value is None
            else str(value_ws.cell(row=r, column=c).value)
            for c in range(1, 7)
        ]
        for r in (1, 2)
    ]
    value_wb.close()

    style_wb = load_workbook(template_path, data_only=False)
    style_ws = style_wb[style_wb.sheetnames[0]]
    styles = {}
    for r in (1, 2, 3):
        for c in range(1, 7):
            cell = style_ws.cell(row=r, column=c)
            styles[(r, c)] = {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
    widths = {col: style_ws.column_dimensions[col].width for col in "ABCDEF"}
    widths["A"] = max(widths["A"] or 0, 40)
    heights = {r: style_ws.row_dimensions[r].height for r in (1, 2, 3)}
    snapshot = {
        "headers": headers,
        "styles": styles,
        "widths": widths,
        "heights": heights,
        "freeze_panes": style_ws.freeze_panes,
        "sheet_title": style_ws.title,
    }
    style_wb.close()
    return snapshot


def create_xlsx(template: dict[str, Any], group_name: str, rows: list[list[str]], output_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = template["sheet_title"]
    worksheet.freeze_panes = template["freeze_panes"]
    worksheet.sheet_format.defaultRowHeight = 15

    for column, width in template["widths"].items():
        worksheet.column_dimensions[column].width = width
    for row_number, height in template["heights"].items():
        if height is not None:
            worksheet.row_dimensions[row_number].height = height

    def apply_style(template_row: int, column_number: int, cell) -> None:
        style = template["styles"][(template_row, column_number)]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
        cell.number_format = "@"
        cell.quotePrefix = False

    for r, header_row in enumerate(template["headers"], start=1):
        for c, value in enumerate(header_row, start=1):
            cell = worksheet.cell(row=r, column=c, value=clean_text_value(value))
            apply_style(r, c, cell)

    label_rows = build_target_rows(group_name, rows)
    for r, label_row in enumerate(label_rows, start=3):
        if template["heights"][3] is not None:
            worksheet.row_dimensions[r].height = template["heights"][3]
        for c, value in enumerate(label_row, start=1):
            clean_value = clean_text_value(value)
            cell = worksheet.cell(row=r, column=c, value=clean_value)
            cell.number_format = "@"
            cell.quotePrefix = False
            if clean_value:
                apply_style(3, c, cell)

    worksheet.auto_filter.ref = f"A2:F{2 + len(label_rows)}"
    workbook.save(output_path)
    workbook.close()


def write_values(sheets_api, spreadsheet_id: str, tab: str, values: list[list[str]]) -> None:
    end_row = len(values)
    sheets_api.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{tab.replace(chr(39), chr(39) * 2)}'!A1:F{end_row}",
        valueInputOption="RAW",
        body={"majorDimension": "ROWS", "values": values},
    ).execute()


def create_master(
    sheets_api,
    groups: OrderedDict[str, list[list[str]]],
    group_to_tab: dict[str, str],
    template_headers: list[list[str]],
) -> None:
    existing = get_sheet_properties(sheets_api, MASTER_SPREADSHEET_ID)
    staging_title = "__KAL_LABELING_STAGING__"
    staging = next((p for p in existing if p["title"] == staging_title), None)
    if staging:
        staging_id = staging["sheetId"]
    else:
        response = sheets_api.spreadsheets().batchUpdate(
            spreadsheetId=MASTER_SPREADSHEET_ID,
            body={"requests": [{"addSheet": {"properties": {"title": staging_title}}}]},
        ).execute()
        staging_id = response["replies"][0]["addSheet"]["properties"]["sheetId"]

    existing = get_sheet_properties(sheets_api, MASTER_SPREADSHEET_ID)
    delete_requests = [
        {"deleteSheet": {"sheetId": p["sheetId"]}}
        for p in existing
        if p["sheetId"] != staging_id
    ]
    if delete_requests:
        sheets_api.spreadsheets().batchUpdate(
            spreadsheetId=MASTER_SPREADSHEET_ID,
            body={"requests": delete_requests},
        ).execute()

    add_requests = []
    for group_name, rows in groups.items():
        add_requests.append({
            "addSheet": {
                "properties": {
                    "title": group_to_tab[group_name],
                    "gridProperties": {
                        "rowCount": max(100, 4 + len(rows) * 3),
                        "columnCount": 6,
                        "frozenRowCount": 2,
                    },
                }
            }
        })
    sheets_api.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        body={"requests": add_requests},
    ).execute()

    properties = get_sheet_properties(sheets_api, MASTER_SPREADSHEET_ID)
    title_to_id = {p["title"]: p["sheetId"] for p in properties}

    for group_name, rows in groups.items():
        tab = group_to_tab[group_name]
        sheet_id = title_to_id[tab]
        values = template_headers + build_target_rows(group_name, rows)
        write_values(sheets_api, MASTER_SPREADSHEET_ID, tab, values)
        last_row = len(values)
        requests = [
            {"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": last_row, "startColumnIndex": 0, "endColumnIndex": 6}, "cell": {"userEnteredFormat": {"numberFormat": {"type": "TEXT"}, "textFormat": {"fontFamily": "Aptos Narrow", "fontSize": 11}}}, "fields": "userEnteredFormat(numberFormat,textFormat.fontFamily,textFormat.fontSize)"}},
            {"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": 6}, "cell": {"userEnteredFormat": {"backgroundColor": rgb("#D8D8D8"), "textFormat": {"bold": True, "foregroundColor": rgb("#000000")}, "borders": {edge: {"style": "SOLID", "color": rgb("#000000")} for edge in ("top", "bottom", "left", "right")}}}, "fields": "userEnteredFormat(backgroundColor,textFormat.bold,textFormat.foregroundColor,borders)"}},
            {"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": 6}, "cell": {"userEnteredFormat": {"wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE"}}, "fields": "userEnteredFormat(wrapStrategy,verticalAlignment)"}},
            {"addConditionalFormatRule": {"rule": {"ranges": [{"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": last_row, "startColumnIndex": 0, "endColumnIndex": 6}], "booleanRule": {"condition": {"type": "CUSTOM_FORMULA", "values": [{"userEnteredValue": "=LEN(A3)>0"}]}, "format": {"backgroundColor": rgb("#FFC000"), "textFormat": {"foregroundColor": rgb("#FF0000")}}}}, "index": 0}},
        ]
        for start, end, size in ((0, 1, 330), (1, 2, 180), (2, 3, 380), (3, 5, 220), (5, 6, 180)):
            requests.append({"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": start, "endIndex": end}, "properties": {"pixelSize": size}, "fields": "pixelSize"}})
        sheets_api.spreadsheets().batchUpdate(
            spreadsheetId=MASTER_SPREADSHEET_ID,
            body={"requests": requests},
        ).execute()

    sheets_api.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        body={"requests": [{"deleteSheet": {"sheetId": staging_id}}]},
    ).execute()


def create_validation_tab(
    sheets_api,
    groups: OrderedDict[str, list[list[str]]],
    group_to_tab: dict[str, str],
    source_row_number_by_id: dict[int, int],
) -> int:
    tab = "00_Datenvalidierung"
    headers = ["Status", "Source-Zeile", "AA-Gruppe", "Master-Reiter", "Source-Feld", "Source-Spalte", "Vorhandener Wert", "Betroffene Labelzeile", "Target-Spalte", "Hinweis", "Source-Link"]
    rules = [("EQM Nummer", "Z", COL_EQM_NUMBER, "Equipment", "A"), ("Messstellen-Beschreibung", "G", COL_DESCRIPTION, "Equipment + Functional Location", "C"), ("Übersetzung auf neue FLO", "X", COL_NEW_FLO, "Functional Location", "A"), ("Asset ID / alt", "E", COL_ASSET_ID, "Legacy-Bezeichnung", "A")]
    issues = []
    for group_name, rows in groups.items():
        for row in rows:
            row_number = source_row_number_by_id[id(row)]
            source_link = f"https://docs.google.com/spreadsheets/d/{SOURCE_SPREADSHEET_ID}/edit#gid={SOURCE_SHEET_GID}&range=A{row_number}:AA{row_number}"
            for field, source_col, index, labels, target_col in rules:
                value = clean_text_value(row[index])
                if value and not PLACEHOLDER_PATTERN.fullmatch(value):
                    continue
                issues.append(["FEHLT" if not value else "PRÜFEN", str(row_number), group_name, group_to_tab[group_name], field, source_col, value, labels, target_col, "Source-Zelle ist leer." if not value else "Source enthält einen Fehler- oder Platzhalterwert.", source_link])
    issues.sort(key=lambda row: (int(row[1]), row[5]))
    values = [headers] + issues if issues else [headers, ["OK", "", "", "", "", "", "", "", "", "Keine fehlenden oder offensichtlichen Fehlerwerte in E, G, X und Z gefunden.", ""]]

    existing = get_sheet_properties(sheets_api, MASTER_SPREADSHEET_ID)
    old = next((p for p in existing if p["title"] == tab), None)
    if old:
        sheets_api.spreadsheets().batchUpdate(spreadsheetId=MASTER_SPREADSHEET_ID, body={"requests": [{"deleteSheet": {"sheetId": old["sheetId"]}}]}).execute()
    added = sheets_api.spreadsheets().batchUpdate(spreadsheetId=MASTER_SPREADSHEET_ID, body={"requests": [{"addSheet": {"properties": {"title": tab, "index": 0, "gridProperties": {"rowCount": max(100, len(values)), "columnCount": len(headers), "frozenRowCount": 1}}}}]}).execute()
    sheet_id = added["replies"][0]["addSheet"]["properties"]["sheetId"]
    sheets_api.spreadsheets().values().update(spreadsheetId=MASTER_SPREADSHEET_ID, range=f"'{tab}'!A1:K{len(values)}", valueInputOption="RAW", body={"majorDimension": "ROWS", "values": values}).execute()
    format_requests = [
        {"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": len(values), "startColumnIndex": 0, "endColumnIndex": len(headers)}, "cell": {"userEnteredFormat": {"numberFormat": {"type": "TEXT"}, "verticalAlignment": "MIDDLE"}}, "fields": "userEnteredFormat(numberFormat,verticalAlignment)"}},
        {"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": len(headers)}, "cell": {"userEnteredFormat": {"backgroundColor": rgb("#B91C1C"), "textFormat": {"bold": True, "foregroundColor": rgb("#FFFFFF")}, "wrapStrategy": "WRAP"}}, "fields": "userEnteredFormat(backgroundColor,textFormat.bold,textFormat.foregroundColor,wrapStrategy)"}},
        {"setBasicFilter": {"filter": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": len(values), "startColumnIndex": 0, "endColumnIndex": len(headers)}}}},
    ]
    for index, size in enumerate((90, 105, 260, 260, 230, 100, 210, 260, 105, 300, 420)):
        format_requests.append({"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": index, "endIndex": index + 1}, "properties": {"pixelSize": size}, "fields": "pixelSize"}})
    sheets_api.spreadsheets().batchUpdate(spreadsheetId=MASTER_SPREADSHEET_ID, body={"requests": format_requests}).execute()
    print(f"Datenvalidierung: {len(issues)} Hinweis(e). https://docs.google.com/spreadsheets/d/{MASTER_SPREADSHEET_ID}/edit#gid={sheet_id}")
    return len(issues)


def upload_or_update(drive_api, local_path: Path, filename: str) -> str:
    from googleapiclient.http import MediaFileUpload

    query = f"'{TARGET_DRIVE_FOLDER_ID}' in parents and name = '{filename.replace(chr(39), chr(92) + chr(39))}' and trashed = false"
    existing = drive_api.files().list(q=query, spaces="drive", fields="files(id)", pageSize=10).execute().get("files", [])
    media = MediaFileUpload(str(local_path), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resumable=True)
    if existing:
        result = drive_api.files().update(fileId=existing[0]["id"], body={"name": filename}, media_body=media, fields="webViewLink").execute()
        return result.get("webViewLink", f"https://drive.google.com/open?id={existing[0]['id']}")
    result = drive_api.files().create(body={"name": filename, "parents": [TARGET_DRIVE_FOLDER_ID]}, media_body=media, fields="webViewLink").execute()
    return result.get("webViewLink", f"https://drive.google.com/open?id={result['id']}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=Path("Vorlage Importdatei.xlsx"))
    parser.add_argument("--credentials-file", type=Path, default=None, help="Optionaler Google-Service-Account-Schlüssel.")
    parser.add_argument("--export", action="store_true", help="XLSX-Dateien ohne Rückfrage erzeugen und in Drive aktualisieren.")
    parser.add_argument("--output-dir", type=Path, default=Path("label_exports"))
    args = parser.parse_args()
    if not args.template.exists():
        raise FileNotFoundError(f"Vorlage nicht gefunden: {args.template}")

    from googleapiclient.discovery import build

    credentials = get_credentials(args.credentials_file)
    sheets_api = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    drive_api = build("drive", "v3", credentials=credentials, cache_discovery=False)
    template = load_template(args.template)

    response = sheets_api.spreadsheets().values().get(spreadsheetId=SOURCE_SPREADSHEET_ID, range=f"'{SOURCE_SHEET_NAME}'!A:AA", valueRenderOption="FORMATTED_VALUE", dateTimeRenderOption="FORMATTED_STRING", majorDimension="ROWS").execute()
    all_rows = response.get("values", [])
    if not all_rows:
        raise RuntimeError("Keine Source-Werte gelesen.")
    source_rows = [normalize_row(row) for row in all_rows[1:]]
    source_row_number_by_id = {id(row): number for number, row in enumerate(source_rows, start=2)}
    filtered = [row for row in source_rows if GROUP_PATTERN.fullmatch(row[COL_GROUP].strip())]
    groups: OrderedDict[str, list[list[str]]] = OrderedDict()
    for row in filtered:
        groups.setdefault(row[COL_GROUP].strip(), []).append(row)
    groups = OrderedDict(sorted(groups.items(), key=lambda item: group_sort_key(item[0])))
    used_titles: set[str] = set()
    group_to_tab = {name: safe_sheet_title(name, used_titles) for name in groups}

    create_master(sheets_api, groups, group_to_tab, template["headers"])
    create_validation_tab(sheets_api, groups, group_to_tab, source_row_number_by_id)
    print(f"Master-Target erstellt: {len(groups)} Reiter, {len(filtered)} Source-Zeilen.")

    if args.export or input("XLSX-Dateien jetzt in Drive schreiben? [ja/NEIN]: ").strip().casefold() in {"ja", "j", "yes", "y"}:
        if args.output_dir.exists():
            shutil.rmtree(args.output_dir)
        args.output_dir.mkdir(parents=True, exist_ok=True)
        for group_name, rows in groups.items():
            filename = safe_filename(group_name)
            local_path = args.output_dir / filename
            create_xlsx(template, group_name, rows, local_path)
            try:
                link = upload_or_update(drive_api, local_path, filename)
                print(f"{filename}: {link}")
            except Exception as exc:
                print(
                    f"WARNUNG: {filename} lokal erstellt, "
                    f"Drive-Upload fehlgeschlagen: {type(exc).__name__}: {exc}"
                )
        print(f"Lokale XLSX-Dateien: {args.output_dir.resolve()}")
    else:
        print("Kein XLSX-Export ausgeführt.")


if __name__ == "__main__":
    main()
